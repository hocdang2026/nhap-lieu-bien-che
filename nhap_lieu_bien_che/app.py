import os, re, sqlite3, secrets, unicodedata, csv
from io import BytesIO, StringIO
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort, session, Response
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get('DATA_DIR', BASE_DIR / 'data'))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / 'app.db'
TEMPLATE_PATH = Path(os.environ.get('EXCEL_TEMPLATE', DATA_DIR / 'template.xlsx'))
CURRENT_XLSX = DATA_DIR / 'current_output.xlsx'
SHEET_NAME = 'CBCC'

FIELD_DEFS = [
    ('temp_quota', 'Biên chế tạm giao năm 2026', 3, 'number'),
    ('present_count', 'Biên chế có mặt tính đến ngày 15/6/2026', 4, 'number'),
    ('assigned_quota', 'Biên chế giao năm 2026', 5, 'number'),
    ('note', 'Ghi chú', 8, 'text'),
]

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.environ.get('SECRET_KEY', 'CHANGE-ME-' + secrets.token_hex(16)),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
    MAX_CONTENT_LENGTH=2 * 1024 * 1024,
)
if os.environ.get('APP_ENV') == 'production':
    app.config['SESSION_COOKIE_SECURE'] = True


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


def slugify(text):
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = text.lower().replace('đ', 'd')
    text = re.sub(r'[^a-z0-9]+', '', text)
    return text[:28] or 'donvi'


def unit_rows_from_excel():
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f'Không tìm thấy sheet {SHEET_NAME}')
    ws = wb[SHEET_NAME]
    units = []
    for row in range(6, ws.max_row + 1):
        stt = ws.cell(row, 1).value
        name = ws.cell(row, 2).value
        if isinstance(stt, (int, float)) and name:
            units.append((row, str(name).strip()))
        elif isinstance(stt, str) and stt.strip().isdigit() and name:
            units.append((row, str(name).strip()))
    return units


def init_db():
    conn = db()
    conn.executescript('''
    CREATE TABLE IF NOT EXISTS departments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        excel_row INTEGER NOT NULL UNIQUE,
        active INTEGER NOT NULL DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL CHECK(role IN ('admin','department')),
        department_id INTEGER,
        must_change_password INTEGER NOT NULL DEFAULT 1,
        failed_attempts INTEGER NOT NULL DEFAULT 0,
        locked_until TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL,
        FOREIGN KEY(department_id) REFERENCES departments(id)
    );
    CREATE TABLE IF NOT EXISTS submissions (
        department_id INTEGER PRIMARY KEY,
        temp_quota INTEGER,
        present_count INTEGER,
        assigned_quota INTEGER,
        note TEXT,
        updated_at TEXT NOT NULL,
        updated_by INTEGER NOT NULL,
        FOREIGN KEY(department_id) REFERENCES departments(id),
        FOREIGN KEY(updated_by) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        department_id INTEGER,
        action TEXT NOT NULL,
        old_data TEXT,
        new_data TEXT,
        ip_address TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(department_id) REFERENCES departments(id)
    );
    CREATE TABLE IF NOT EXISTS field_settings (
        field_key TEXT PRIMARY KEY,
        enabled INTEGER NOT NULL DEFAULT 1,
        required INTEGER NOT NULL DEFAULT 0
    );
    ''')

    # Seed departments from workbook.
    units = unit_rows_from_excel()
    for row, name in units:
        conn.execute('INSERT OR IGNORE INTO departments(name, excel_row) VALUES (?,?)', (name, row))
    for key, _label, _col, _kind in FIELD_DEFS:
        conn.execute('INSERT OR IGNORE INTO field_settings(field_key, enabled, required) VALUES (?,?,?)', (key, 1, 0))
    conn.commit()

    # Seed admin.
    existing = conn.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
    if not existing:
        admin_password = os.environ.get('ADMIN_INITIAL_PASSWORD', 'Admin@2026!')
        conn.execute('''INSERT INTO users(username,password_hash,role,must_change_password,created_at)
                        VALUES (?,?,?,?,?)''',
                     ('admin', generate_password_hash(admin_password), 'admin', 1, datetime.now().isoformat()))

    # Seed department accounts from a pre-generated credential file when present.
    seed_path = DATA_DIR / 'initial_accounts_seed.csv'
    seed_map = {}
    if seed_path.exists():
        with open(seed_path, 'r', encoding='utf-8-sig', newline='') as f:
            for row in csv.DictReader(f):
                seed_map[row['Đơn vị'].strip()] = (row['Tên đăng nhập'].strip(), row['Mật khẩu tạm'])

    departments = conn.execute('SELECT * FROM departments ORDER BY excel_row').fetchall()
    used = {r['username'] for r in conn.execute('SELECT username FROM users').fetchall()}
    for d in departments:
        if conn.execute('SELECT id FROM users WHERE department_id=?', (d['id'],)).fetchone():
            continue
        if d['name'] in seed_map:
            username, password = seed_map[d['name']]
        else:
            base = slugify(d['name'])
            username = base
            n = 2
            while username in used:
                username = f'{base}{n}'
                n += 1
            password = secrets.token_urlsafe(7) + '!6A'
        conn.execute('''INSERT INTO users(username,password_hash,role,department_id,must_change_password,created_at)
                        VALUES (?,?,?,?,?,?)''',
                     (username, generate_password_hash(password), 'department', d['id'], 1, datetime.now().isoformat()))
        used.add(username)
    conn.commit()
    conn.close()
    rebuild_current_excel()


def current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    conn = db()
    row = conn.execute('''SELECT u.*, d.name AS department_name, d.excel_row
                          FROM users u LEFT JOIN departments d ON d.id=u.department_id
                          WHERE u.id=? AND u.active=1''', (uid,)).fetchone()
    conn.close()
    return row


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for('login', next=request.path))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u:
            return redirect(url_for('login'))
        if u['role'] != 'admin':
            abort(403)
        return fn(*args, **kwargs)
    return wrapper


def get_field_settings(conn=None):
    own = conn is None
    conn = conn or db()
    rows = {r['field_key']: r for r in conn.execute('SELECT * FROM field_settings').fetchall()}
    out = []
    for key, label, col, kind in FIELD_DEFS:
        r = rows.get(key)
        out.append({
            'key': key, 'label': label, 'column': col, 'kind': kind,
            'enabled': bool(r['enabled']) if r else True,
            'required': bool(r['required']) if r else False,
        })
    if own:
        conn.close()
    return out


def safe_int(value, label):
    value = (value or '').strip()
    if value == '':
        return None
    if not re.fullmatch(r'\d+', value):
        raise ValueError(f'{label} phải là số nguyên không âm.')
    n = int(value)
    if n > 100000:
        raise ValueError(f'{label} quá lớn, vui lòng kiểm tra lại.')
    return n


def rebuild_current_excel(output_path=CURRENT_XLSX):
    if not TEMPLATE_PATH.exists():
        return
    conn = db() if DB_PATH.exists() else None
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb[SHEET_NAME]
    if conn:
        rows = conn.execute('''SELECT d.excel_row, s.temp_quota, s.present_count, s.assigned_quota, s.note
                               FROM departments d LEFT JOIN submissions s ON s.department_id=d.id
                               ORDER BY d.excel_row''').fetchall()
        for r in rows:
            er = r['excel_row']
            ws.cell(er, 3).value = r['temp_quota']
            ws.cell(er, 4).value = r['present_count']
            ws.cell(er, 5).value = r['assigned_quota']
            ws.cell(er, 6).value = f'=E{er}-C{er}'
            ws.cell(er, 7).value = f'=E{er}-D{er}'
            ws.cell(er, 8).value = r['note']
        conn.close()
    wb.save(output_path)


def csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


@app.before_request
def csrf_protect():
    if request.method == 'POST':
        sent = request.form.get('_csrf_token', '')
        expected = session.get('_csrf_token', '')
        if not sent or not expected or not secrets.compare_digest(sent, expected):
            abort(400, description='Phiên làm việc không hợp lệ. Vui lòng tải lại trang và thử lại.')


@app.context_processor
def inject_user():
    return {'me': current_user(), 'csrf_token': csrf_token}


@app.route('/health')
def health():
    return {'status': 'ok'}


@app.route('/', methods=['GET'])
def home():
    if current_user():
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip().lower()
        password = request.form.get('password') or ''
        conn = db()
        u = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        now = datetime.now()
        if not u or not u['active']:
            conn.close(); flash('Sai tên đăng nhập hoặc mật khẩu.', 'danger'); return render_template('login.html')
        if u['locked_until']:
            try:
                locked_until = datetime.fromisoformat(u['locked_until'])
                if locked_until > now:
                    conn.close(); flash('Tài khoản đang tạm khóa do nhập sai nhiều lần. Vui lòng thử lại sau.', 'danger'); return render_template('login.html')
            except ValueError:
                pass
        if not check_password_hash(u['password_hash'], password):
            attempts = u['failed_attempts'] + 1
            locked = (now + timedelta(minutes=15)).isoformat() if attempts >= 5 else None
            conn.execute('UPDATE users SET failed_attempts=?, locked_until=? WHERE id=?', (0 if locked else attempts, locked, u['id']))
            conn.commit(); conn.close()
            flash('Sai tên đăng nhập hoặc mật khẩu.', 'danger'); return render_template('login.html')
        conn.execute('UPDATE users SET failed_attempts=0, locked_until=NULL WHERE id=?', (u['id'],))
        conn.commit(); conn.close()
        session.clear(); session['user_id'] = u['id']; session.permanent = True
        if u['must_change_password']:
            return redirect(url_for('change_password'))
        return redirect(url_for('dashboard'))
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    u = current_user()
    if request.method == 'POST':
        old = request.form.get('old_password') or ''
        new = request.form.get('new_password') or ''
        confirm = request.form.get('confirm_password') or ''
        if not check_password_hash(u['password_hash'], old):
            flash('Mật khẩu hiện tại không đúng.', 'danger')
        elif len(new) < 8 or not re.search(r'[A-Za-z]', new) or not re.search(r'\d', new):
            flash('Mật khẩu mới phải từ 8 ký tự và có cả chữ lẫn số.', 'danger')
        elif new != confirm:
            flash('Xác nhận mật khẩu chưa khớp.', 'danger')
        else:
            conn = db(); conn.execute('UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?', (generate_password_hash(new), u['id']))
            conn.commit(); conn.close(); flash('Đổi mật khẩu thành công.', 'success')
            return redirect(url_for('dashboard'))
    return render_template('change_password.html')


@app.route('/dashboard')
@login_required
def dashboard():
    u = current_user()
    if u['must_change_password']:
        return redirect(url_for('change_password'))
    if u['role'] == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('department_form'))


@app.route('/nhap-lieu', methods=['GET', 'POST'])
@login_required
def department_form():
    u = current_user()
    if u['role'] != 'department':
        return redirect(url_for('admin_dashboard'))
    if u['must_change_password']:
        return redirect(url_for('change_password'))
    conn = db()
    sub = conn.execute('SELECT * FROM submissions WHERE department_id=?', (u['department_id'],)).fetchone()
    fields = get_field_settings(conn)
    if request.method == 'POST':
        values = {
            'temp_quota': sub['temp_quota'] if sub else None,
            'present_count': sub['present_count'] if sub else None,
            'assigned_quota': sub['assigned_quota'] if sub else None,
            'note': (sub['note'] or '') if sub else '',
        }
        try:
            for f in fields:
                if not f['enabled']:
                    continue
                raw = request.form.get(f['key'])
                if f['required'] and (raw is None or str(raw).strip() == ''):
                    raise ValueError(f"{f['label']} là dữ liệu bắt buộc phải nhập.")
                if f['kind'] == 'number':
                    values[f['key']] = safe_int(raw, f['label'])
                else:
                    values[f['key']] = (raw or '').strip()[:1000]
        except ValueError as e:
            conn.close(); flash(str(e), 'danger'); return render_template('department_form.html', sub=sub, fields=fields)
        old_data = dict(sub) if sub else None
        now = datetime.now().isoformat(timespec='seconds')
        conn.execute('''INSERT INTO submissions(department_id,temp_quota,present_count,assigned_quota,note,updated_at,updated_by)
                        VALUES (?,?,?,?,?,?,?)
                        ON CONFLICT(department_id) DO UPDATE SET temp_quota=excluded.temp_quota,
                        present_count=excluded.present_count, assigned_quota=excluded.assigned_quota,
                        note=excluded.note, updated_at=excluded.updated_at, updated_by=excluded.updated_by''',
                     (u['department_id'], values['temp_quota'], values['present_count'], values['assigned_quota'], values['note'], now, u['id']))
        new_data = f"C={values['temp_quota']}; D={values['present_count']}; E={values['assigned_quota']}; H={values['note']}"
        conn.execute('''INSERT INTO audit_logs(user_id,department_id,action,old_data,new_data,ip_address,created_at)
                        VALUES (?,?,?,?,?,?,?)''',
                     (u['id'], u['department_id'], 'Cập nhật số liệu', str(old_data), new_data, request.remote_addr, now))
        conn.commit(); conn.close()
        rebuild_current_excel()
        flash('Đã lưu dữ liệu thành công. Dữ liệu đã được cập nhật vào file Excel tổng.', 'success')
        return redirect(url_for('department_form'))
    conn.close()
    return render_template('department_form.html', sub=sub, fields=fields)


@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = db()
    units = conn.execute('''SELECT d.*, u.username, u.active AS user_active, u.id AS user_id,
                           s.updated_at, s.temp_quota, s.present_count, s.assigned_quota
                           FROM departments d
                           LEFT JOIN users u ON u.department_id=d.id
                           LEFT JOIN submissions s ON s.department_id=d.id
                           ORDER BY d.excel_row''').fetchall()
    stats = conn.execute('''SELECT (SELECT COUNT(*) FROM departments) total,
                           (SELECT COUNT(*) FROM submissions) submitted''').fetchone()
    fields = get_field_settings(conn)
    conn.close()
    return render_template('admin.html', units=units, stats=stats, fields=fields)


@app.route('/admin/field-settings', methods=['POST'])
@admin_required
def update_field_settings():
    conn = db()
    enabled_count = 0
    for key, label, _col, _kind in FIELD_DEFS:
        enabled = 1 if request.form.get(f'enabled_{key}') == '1' else 0
        required = 1 if enabled and request.form.get(f'required_{key}') == '1' else 0
        enabled_count += enabled
        conn.execute('''INSERT INTO field_settings(field_key,enabled,required) VALUES (?,?,?)
                        ON CONFLICT(field_key) DO UPDATE SET enabled=excluded.enabled, required=excluded.required''',
                     (key, enabled, required))
    if enabled_count == 0:
        conn.rollback(); conn.close()
        flash('Phải chọn ít nhất 1 trường cho đơn vị nhập.', 'danger')
        return redirect(url_for('admin_dashboard'))
    now = datetime.now().isoformat(timespec='seconds')
    u = current_user()
    config_text = '; '.join(
        f"{f['label']}: {'Bật' if request.form.get('enabled_'+f['key']) == '1' else 'Tắt'}"
        for f in get_field_settings(conn)
    )
    conn.execute('''INSERT INTO audit_logs(user_id,department_id,action,old_data,new_data,ip_address,created_at)
                    VALUES (?,?,?,?,?,?,?)''',
                 (u['id'], None, 'Cấu hình trường nhập liệu', None, config_text, request.remote_addr, now))
    conn.commit(); conn.close()
    flash('Đã cập nhật các trường dữ liệu cần nhập. Form của tất cả đơn vị đã thay đổi theo cấu hình mới.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/unit/<int:department_id>')
@admin_required
def admin_unit(department_id):
    conn = db()
    unit = conn.execute('''SELECT d.*, u.username, u.id AS user_id, u.active AS user_active,
                          s.* FROM departments d LEFT JOIN users u ON u.department_id=d.id
                          LEFT JOIN submissions s ON s.department_id=d.id WHERE d.id=?''', (department_id,)).fetchone()
    logs = conn.execute('''SELECT a.*, u.username FROM audit_logs a LEFT JOIN users u ON u.id=a.user_id
                           WHERE a.department_id=? ORDER BY a.id DESC LIMIT 50''', (department_id,)).fetchall()
    conn.close()
    if not unit: abort(404)
    return render_template('admin_unit.html', unit=unit, logs=logs)


@app.route('/admin/reset-password/<int:user_id>', methods=['POST'])
@admin_required
def reset_password(user_id):
    new_password = secrets.token_urlsafe(7) + '!6A'
    conn = db()
    u = conn.execute("SELECT u.*, d.name AS department_name FROM users u LEFT JOIN departments d ON d.id=u.department_id WHERE u.id=? AND u.role='department'", (user_id,)).fetchone()
    if not u:
        conn.close(); abort(404)
    conn.execute('UPDATE users SET password_hash=?, must_change_password=1, failed_attempts=0, locked_until=NULL WHERE id=?', (generate_password_hash(new_password), user_id))
    conn.commit(); conn.close()
    flash(f'Mật khẩu tạm mới của {u["department_name"]}: {new_password}', 'warning')
    return redirect(url_for('admin_unit', department_id=u['department_id']))


@app.route('/admin/toggle-user/<int:user_id>', methods=['POST'])
@admin_required
def toggle_user(user_id):
    conn = db()
    u = conn.execute("SELECT * FROM users WHERE id=? AND role='department'", (user_id,)).fetchone()
    if not u: conn.close(); abort(404)
    conn.execute('UPDATE users SET active=? WHERE id=?', (0 if u['active'] else 1, user_id))
    conn.commit(); conn.close(); flash('Đã cập nhật trạng thái tài khoản.', 'success')
    return redirect(url_for('admin_unit', department_id=u['department_id']))


@app.route('/admin/credentials.csv')
@admin_required
def download_credentials():
    conn = db()
    rows = conn.execute('''SELECT d.name, u.username, u.must_change_password, u.active
                           FROM departments d JOIN users u ON u.department_id=d.id ORDER BY d.excel_row''').fetchall()
    conn.close()
    sio = StringIO(); w = csv.writer(sio)
    w.writerow(['Đơn vị','Tên đăng nhập','Mật khẩu','Ghi chú'])
    for r in rows:
        w.writerow([r['name'], r['username'], '(Không thể xem lại - hãy Đặt lại mật khẩu nếu quên)', 'Hoạt động' if r['active'] else 'Đã khóa'])
    data = BytesIO(sio.getvalue().encode('utf-8-sig'))
    return send_file(data, mimetype='text/csv', as_attachment=True, download_name='danh_sach_tai_khoan.csv')


@app.route('/admin/export.xlsx')
@admin_required
def export_excel():
    rebuild_current_excel()
    return send_file(CURRENT_XLSX, as_attachment=True, download_name='THUYET_MINH_BIEN_CHE_2026_TONG_HOP.xlsx')


@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', title='Không có quyền truy cập', message='Bạn không có quyền truy cập chức năng này.'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', title='Không tìm thấy', message='Trang bạn yêu cầu không tồn tại.'), 404


init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=os.environ.get('FLASK_DEBUG') == '1')
