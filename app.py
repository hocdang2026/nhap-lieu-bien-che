import os, re, sqlite3, secrets, unicodedata, csv
from io import BytesIO, StringIO
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort, session, Response
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get('DATA_DIR', BASE_DIR / 'data'))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / 'app.db'
TEMPLATE_PATH = Path(os.environ.get('EXCEL_TEMPLATE', DATA_DIR / 'template.xlsx'))
CURRENT_XLSX = DATA_DIR / 'current_output.xlsx'
UPLOAD_DIR = DATA_DIR / 'uploads'
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
SHEET_NAME = 'CBCC'

FIELD_DEFS = [
    ('temp_quota', 'Biên chế tạm giao năm 2026', 3, 'number'),
    ('present_count', 'Biên chế có mặt tính đến ngày 15/6/2026', 4, 'number'),
    ('assigned_quota', 'Biên chế giao năm 2026', 5, 'number'),
    ('note', 'Ghi chú', 8, 'text'),
]

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.environ.get('SECRET_KEY', 'CHANGE-ME-' + secrets.token_hex(16)),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
    MAX_CONTENT_LENGTH=15 * 1024 * 1024,
)
if os.environ.get('APP_ENV') == 'production':
    app.config['SESSION_COOKIE_SECURE'] = True


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


def slugify(text):
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = text.lower().replace('đ', 'd')
    text = re.sub(r'[^a-z0-9]+', '', text)
    return text[:28] or 'donvi'



def norm_text(value):
    if value is None:
        return ''
    text = unicodedata.normalize('NFD', str(value))
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = text.lower().replace('đ', 'd')
    return re.sub(r'\s+', ' ', text).strip()


def convert_xls_to_xlsx(src_path, dest_path):
    """Chuyển .xls cũ sang .xlsx theo dữ liệu ô. Định dạng/formula phức tạp có thể không giữ nguyên."""
    import xlrd
    book = xlrd.open_workbook(src_path)
    out = Workbook()
    out.remove(out.active)
    for sh in book.sheets():
        ws = out.create_sheet(title=(sh.name[:31] or 'Sheet'))
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, book.datemode)
                    except Exception:
                        pass
                ws.cell(r + 1, c + 1).value = val
    out.save(dest_path)


def inspect_excel(path):
    """Tự nhận dạng sheet/dòng tiêu đề/cột đơn vị và trả preview để Admin xác nhận."""
    wb = load_workbook(path, data_only=False, read_only=True)
    result = []
    unit_words = ('ten co quan', 'ten don vi', 'don vi', 'phong ban', 'co quan, to chuc', 'co quan', 'to chuc')
    for ws in wb.worksheets:
        max_col = min(ws.max_column or 1, 80)
        max_row = min(ws.max_row or 1, 60)
        best_row, best_score = 1, -1
        for r in range(1, max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            nonempty = [v for v in vals if v not in (None, '')]
            strings = [v for v in nonempty if isinstance(v, str)]
            score = len(nonempty) * 3 + len(strings)
            if score > best_score:
                best_row, best_score = r, score
        headers = []
        for c in range(1, max_col + 1):
            v = ws.cell(best_row, c).value
            headers.append(str(v).strip() if v not in (None, '') else '')
        unit_col = None
        unit_score = -1
        for idx, h in enumerate(headers, start=1):
            nh = norm_text(h)
            score = 0
            for pos, word in enumerate(unit_words):
                if word and word in nh:
                    score = 100 - pos
                    break
            if score > unit_score:
                unit_col, unit_score = idx, score
        if unit_score <= 0:
            counts = []
            for c in range(1, max_col + 1):
                cnt = 0
                for r in range(best_row + 1, min(ws.max_row, best_row + 120) + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.strip():
                        cnt += 1
                counts.append(cnt)
            unit_col = (counts.index(max(counts)) + 1) if counts else 1
        preview = []
        for r in range(best_row + 1, min(ws.max_row, best_row + 12) + 1):
            preview.append([ws.cell(r, c).value for c in range(1, min(max_col, 12) + 1)])
        result.append({'sheet': ws.title, 'header_row': best_row, 'unit_col': unit_col,
                       'headers': headers, 'preview': preview, 'max_row': ws.max_row, 'max_col': ws.max_column})
    return result


def get_template_meta(conn=None):
    own = conn is None
    conn = conn or db()
    rows = conn.execute('SELECT key,value FROM template_meta').fetchall()
    meta = {r['key']: r['value'] for r in rows}
    if own:
        conn.close()
    return meta


def active_template_path(conn=None):
    try:
        meta = get_template_meta(conn)
        p = meta.get('template_path')
        if p and Path(p).exists():
            return Path(p)
    except Exception:
        pass
    return TEMPLATE_PATH


def set_template_meta(conn, **kwargs):
    for k, v in kwargs.items():
        conn.execute("""INSERT INTO template_meta(key,value) VALUES (?,?)
                        ON CONFLICT(key) DO UPDATE SET value=excluded.value""", (k, '' if v is None else str(v)))


def ensure_department_user(conn, department_id, department_name):
    if conn.execute('SELECT id FROM users WHERE department_id=?', (department_id,)).fetchone():
        return None
    used = {r['username'] for r in conn.execute('SELECT username FROM users').fetchall()}
    base = slugify(department_name)
    username = base
    n = 2
    while username in used:
        username = f'{base}{n}'
        n += 1
    password = secrets.token_urlsafe(7) + '!6A'
    conn.execute("""INSERT INTO users(username,password_hash,role,department_id,must_change_password,created_at)
                    VALUES (?,?,?,?,?,?)""",
                 (username, generate_password_hash(password), 'department', department_id, 1, datetime.now().isoformat()))
    return username, password

def unit_rows_from_excel():
    meta = {}
    if DB_PATH.exists():
        try:
            conn = db()
            if conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='template_meta'").fetchone():
                meta = get_template_meta(conn)
            conn.close()
        except Exception:
            meta = {}
    sheet_name = meta.get('sheet_name') or SHEET_NAME
    unit_col = int(meta.get('unit_col') or 2)
    header_row = int(meta.get('header_row') or 3)
    template_path = active_template_path()
    wb = load_workbook(template_path, data_only=False, keep_vba=template_path.suffix.lower()=='.xlsm')
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f'Không tìm thấy sheet {sheet_name}')
    ws = wb[sheet_name]
    units = []
    if meta.get('sheet_name'):
        for row in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row, unit_col).value
            if name is None or str(name).strip() == '':
                continue
            text = str(name).strip()
            # bỏ các dòng tổng/tiêu đề nhóm thường gặp; Admin vẫn được xác nhận lại khi upload.
            n = norm_text(text)
            if n.startswith('tong') or n in ('cap tinh', 'cap xa', 'du phong'):
                continue
            units.append((row, text))
    else:
        for row in range(6, ws.max_row + 1):
            stt = ws.cell(row, 1).value
            name = ws.cell(row, 2).value
            if isinstance(stt, (int, float)) and name:
                units.append((row, str(name).strip()))
            elif isinstance(stt, str) and stt.strip().isdigit() and name:
                units.append((row, str(name).strip()))
    return units


def init_db():
    conn = db()
    conn.executescript('''
    CREATE TABLE IF NOT EXISTS departments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        excel_row INTEGER NOT NULL UNIQUE,
        active INTEGER NOT NULL DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL CHECK(role IN ('admin','department')),
        department_id INTEGER,
        must_change_password INTEGER NOT NULL DEFAULT 1,
        failed_attempts INTEGER NOT NULL DEFAULT 0,
        locked_until TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL,
        FOREIGN KEY(department_id) REFERENCES departments(id)
    );
    CREATE TABLE IF NOT EXISTS submissions (
        department_id INTEGER PRIMARY KEY,
        temp_quota INTEGER,
        present_count INTEGER,
        assigned_quota INTEGER,
        note TEXT,
        updated_at TEXT NOT NULL,
        updated_by INTEGER NOT NULL,
        FOREIGN KEY(department_id) REFERENCES departments(id),
        FOREIGN KEY(updated_by) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        department_id INTEGER,
        action TEXT NOT NULL,
        old_data TEXT,
        new_data TEXT,
        ip_address TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(department_id) REFERENCES departments(id)
    );
    CREATE TABLE IF NOT EXISTS field_settings (
        field_key TEXT PRIMARY KEY,
        enabled INTEGER NOT NULL DEFAULT 1,
        required INTEGER NOT NULL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS template_meta (
        key TEXT PRIMARY KEY,
        value TEXT
    );
    CREATE TABLE IF NOT EXISTS dynamic_fields (
        field_key TEXT PRIMARY KEY,
        label TEXT NOT NULL,
        excel_column INTEGER NOT NULL,
        kind TEXT NOT NULL DEFAULT 'text',
        enabled INTEGER NOT NULL DEFAULT 1,
        required INTEGER NOT NULL DEFAULT 0,
        sort_order INTEGER NOT NULL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS dynamic_values (
        department_id INTEGER NOT NULL,
        field_key TEXT NOT NULL,
        value_text TEXT,
        updated_at TEXT,
        updated_by INTEGER,
        PRIMARY KEY(department_id, field_key),
        FOREIGN KEY(department_id) REFERENCES departments(id),
        FOREIGN KEY(updated_by) REFERENCES users(id)
    );
    ''')

    # Seed departments from workbook.
    units = unit_rows_from_excel()
    for row, name in units:
        conn.execute('INSERT OR IGNORE INTO departments(name, excel_row) VALUES (?,?)', (name, row))
    for key, _label, _col, _kind in FIELD_DEFS:
        conn.execute('INSERT OR IGNORE INTO field_settings(field_key, enabled, required) VALUES (?,?,?)', (key, 1, 0))
    conn.commit()

    # Seed admin.
    existing = conn.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
    if not existing:
        admin_password = os.environ.get('ADMIN_INITIAL_PASSWORD', 'Admin@2026!')
        conn.execute('''INSERT INTO users(username,password_hash,role,must_change_password,created_at)
                        VALUES (?,?,?,?,?)''',
                     ('admin', generate_password_hash(admin_password), 'admin', 1, datetime.now().isoformat()))

    # Seed department accounts from a pre-generated credential file when present.
    seed_path = DATA_DIR / 'initial_accounts_seed.csv'
    seed_map = {}
    if seed_path.exists():
        with open(seed_path, 'r', encoding='utf-8-sig', newline='') as f:
            for row in csv.DictReader(f):
                seed_map[row['Đơn vị'].strip()] = (row['Tên đăng nhập'].strip(), row['Mật khẩu tạm'])

    departments = conn.execute('SELECT * FROM departments WHERE active=1 ORDER BY excel_row').fetchall()
    used = {r['username'] for r in conn.execute('SELECT username FROM users').fetchall()}
    for d in departments:
        if conn.execute('SELECT id FROM users WHERE department_id=?', (d['id'],)).fetchone():
            continue
        if d['name'] in seed_map:
            username, password = seed_map[d['name']]
        else:
            base = slugify(d['name'])
            username = base
            n = 2
            while username in used:
                username = f'{base}{n}'
                n += 1
            password = secrets.token_urlsafe(7) + '!6A'
        conn.execute('''INSERT INTO users(username,password_hash,role,department_id,must_change_password,created_at)
                        VALUES (?,?,?,?,?,?)''',
                     (username, generate_password_hash(password), 'department', d['id'], 1, datetime.now().isoformat()))
        used.add(username)
    conn.commit()
    conn.close()
    rebuild_current_excel()


def current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    conn = db()
    row = conn.execute('''SELECT u.*, d.name AS department_name, d.excel_row
                          FROM users u LEFT JOIN departments d ON d.id=u.department_id
                          WHERE u.id=? AND u.active=1 AND (u.role='admin' OR d.active=1)''', (uid,)).fetchone()
    conn.close()
    return row


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for('login', next=request.path))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u:
            return redirect(url_for('login'))
        if u['role'] != 'admin':
            abort(403)
        return fn(*args, **kwargs)
    return wrapper


def get_field_settings(conn=None):
    own = conn is None
    conn = conn or db()
    dyn = conn.execute('SELECT * FROM dynamic_fields ORDER BY sort_order, excel_column').fetchall()
    if dyn:
        out = [{
            'key': r['field_key'], 'label': r['label'], 'column': r['excel_column'], 'column_letter': get_column_letter(r['excel_column']), 'kind': r['kind'],
            'enabled': bool(r['enabled']), 'required': bool(r['required'])
        } for r in dyn]
    else:
        rows = {r['field_key']: r for r in conn.execute('SELECT * FROM field_settings').fetchall()}
        out = []
        for key, label, col, kind in FIELD_DEFS:
            r = rows.get(key)
            out.append({
                'key': key, 'label': label, 'column': col, 'column_letter': get_column_letter(col), 'kind': kind,
                'enabled': bool(r['enabled']) if r else True,
                'required': bool(r['required']) if r else False,
            })
    if own:
        conn.close()
    return out


def safe_int(value, label):
    value = (value or '').strip()
    if value == '':
        return None
    if not re.fullmatch(r'\d+', value):
        raise ValueError(f'{label} phải là số nguyên không âm.')
    n = int(value)
    if n > 100000:
        raise ValueError(f'{label} quá lớn, vui lòng kiểm tra lại.')
    return n


def rebuild_current_excel(output_path=CURRENT_XLSX):
    template_path = active_template_path()
    if not template_path.exists():
        return
    conn = db() if DB_PATH.exists() else None
    meta = get_template_meta(conn) if conn else {}
    sheet_name = meta.get('sheet_name') or SHEET_NAME
    wb = load_workbook(template_path, data_only=False, keep_vba=template_path.suffix.lower()=='.xlsm')
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    if conn:
        dyn_fields = conn.execute('SELECT * FROM dynamic_fields ORDER BY sort_order, excel_column').fetchall()
        if dyn_fields:
            rows = conn.execute('SELECT id, excel_row FROM departments WHERE active=1 ORDER BY excel_row').fetchall()
            for d in rows:
                vals = {r['field_key']: r['value_text'] for r in conn.execute(
                    'SELECT field_key,value_text FROM dynamic_values WHERE department_id=?', (d['id'],)).fetchall()}
                for f in dyn_fields:
                    if f['field_key'] not in vals:
                        continue
                    raw = vals[f['field_key']]
                    if f['kind'] == 'number' and raw not in (None, ''):
                        try:
                            value = int(raw)
                        except ValueError:
                            try: value = float(raw)
                            except ValueError: value = raw
                    else:
                        value = raw
                    ws.cell(d['excel_row'], f['excel_column']).value = value
        else:
            rows = conn.execute('''SELECT d.excel_row, s.temp_quota, s.present_count, s.assigned_quota, s.note
                                   FROM departments d LEFT JOIN submissions s ON s.department_id=d.id
                                   ORDER BY d.excel_row''').fetchall()
            for r in rows:
                er = r['excel_row']
                ws.cell(er, 3).value = r['temp_quota']
                ws.cell(er, 4).value = r['present_count']
                ws.cell(er, 5).value = r['assigned_quota']
                ws.cell(er, 6).value = f'=E{er}-C{er}'
                ws.cell(er, 7).value = f'=E{er}-D{er}'
                ws.cell(er, 8).value = r['note']
        conn.close()
    wb.save(output_path)


def csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


@app.before_request
def csrf_protect():
    if request.method == 'POST':
        sent = request.form.get('_csrf_token', '')
        expected = session.get('_csrf_token', '')
        if not sent or not expected or not secrets.compare_digest(sent, expected):
            abort(400, description='Phiên làm việc không hợp lệ. Vui lòng tải lại trang và thử lại.')


@app.context_processor
def inject_user():
    return {'me': current_user(), 'csrf_token': csrf_token}


@app.route('/health')
def health():
    return {'status': 'ok'}


@app.route('/', methods=['GET'])
def home():
    if current_user():
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip().lower()
        password = request.form.get('password') or ''
        conn = db()
        u = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        now = datetime.now()
        if not u or not u['active']:
            conn.close(); flash('Sai tên đăng nhập hoặc mật khẩu.', 'danger'); return render_template('login.html')
        if u['locked_until']:
            try:
                locked_until = datetime.fromisoformat(u['locked_until'])
                if locked_until > now:
                    conn.close(); flash('Tài khoản đang tạm khóa do nhập sai nhiều lần. Vui lòng thử lại sau.', 'danger'); return render_template('login.html')
            except ValueError:
                pass
        if not check_password_hash(u['password_hash'], password):
            attempts = u['failed_attempts'] + 1
            locked = (now + timedelta(minutes=15)).isoformat() if attempts >= 5 else None
            conn.execute('UPDATE users SET failed_attempts=?, locked_until=? WHERE id=?', (0 if locked else attempts, locked, u['id']))
            conn.commit(); conn.close()
            flash('Sai tên đăng nhập hoặc mật khẩu.', 'danger'); return render_template('login.html')
        conn.execute('UPDATE users SET failed_attempts=0, locked_until=NULL WHERE id=?', (u['id'],))
        conn.commit(); conn.close()
        session.clear(); session['user_id'] = u['id']; session.permanent = True
        if u['must_change_password']:
            return redirect(url_for('change_password'))
        return redirect(url_for('dashboard'))
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    u = current_user()
    if request.method == 'POST':
        old = request.form.get('old_password') or ''
        new = request.form.get('new_password') or ''
        confirm = request.form.get('confirm_password') or ''
        if not check_password_hash(u['password_hash'], old):
            flash('Mật khẩu hiện tại không đúng.', 'danger')
        elif len(new) < 8 or not re.search(r'[A-Za-z]', new) or not re.search(r'\d', new):
            flash('Mật khẩu mới phải từ 8 ký tự và có cả chữ lẫn số.', 'danger')
        elif new != confirm:
            flash('Xác nhận mật khẩu chưa khớp.', 'danger')
        else:
            conn = db(); conn.execute('UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?', (generate_password_hash(new), u['id']))
            conn.commit(); conn.close(); flash('Đổi mật khẩu thành công.', 'success')
            return redirect(url_for('dashboard'))
    return render_template('change_password.html')


@app.route('/dashboard')
@login_required
def dashboard():
    u = current_user()
    if u['must_change_password']:
        return redirect(url_for('change_password'))
    if u['role'] == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('department_form'))


@app.route('/nhap-lieu', methods=['GET', 'POST'])
@login_required
def department_form():
    u = current_user()
    if u['role'] != 'department':
        return redirect(url_for('admin_dashboard'))
    if u['must_change_password']:
        return redirect(url_for('change_password'))
    conn = db()
    fields = get_field_settings(conn)
    dyn_mode = bool(conn.execute('SELECT 1 FROM dynamic_fields LIMIT 1').fetchone())
    meta_sub = conn.execute('SELECT * FROM submissions WHERE department_id=?', (u['department_id'],)).fetchone()
    if dyn_mode:
        values = {r['field_key']: r['value_text'] for r in conn.execute(
            'SELECT field_key,value_text FROM dynamic_values WHERE department_id=?', (u['department_id'],)).fetchall()}
        sub = dict(values)
        if meta_sub:
            sub['updated_at'] = meta_sub['updated_at']
    else:
        sub = dict(meta_sub) if meta_sub else None
    if request.method == 'POST':
        old_data = dict(sub) if sub else None
        new_values = {}
        try:
            for f in fields:
                if not f['enabled']:
                    continue
                raw = request.form.get(f['key'])
                if f['required'] and (raw is None or str(raw).strip() == ''):
                    raise ValueError(f"{f['label']} là dữ liệu bắt buộc.")
                if f['kind'] == 'number':
                    val = safe_int(raw, f['label'])
                    new_values[f['key']] = '' if val is None else str(val)
                else:
                    new_values[f['key']] = (raw or '').strip()[:4000]
        except ValueError as e:
            conn.close(); flash(str(e), 'danger')
            temp_sub = dict(sub or {})
            temp_sub.update(new_values)
            return render_template('department_form.html', sub=temp_sub, fields=fields)
        now = datetime.now().isoformat(timespec='seconds')
        if dyn_mode:
            for f in fields:
                if not f['enabled']:
                    continue
                conn.execute('''INSERT INTO dynamic_values(department_id,field_key,value_text,updated_at,updated_by)
                                VALUES (?,?,?,?,?)
                                ON CONFLICT(department_id,field_key) DO UPDATE SET value_text=excluded.value_text,
                                updated_at=excluded.updated_at,updated_by=excluded.updated_by''',
                             (u['department_id'], f['key'], new_values.get(f['key'], ''), now, u['id']))
            # submissions giữ vai trò mốc trạng thái đã nhập/chưa nhập để tương thích dashboard.
            conn.execute('''INSERT INTO submissions(department_id,temp_quota,present_count,assigned_quota,note,updated_at,updated_by)
                            VALUES (?,?,?,?,?,?,?)
                            ON CONFLICT(department_id) DO UPDATE SET updated_at=excluded.updated_at,updated_by=excluded.updated_by''',
                         (u['department_id'], None, None, None, None, now, u['id']))
            new_data = '; '.join(f"{k}={v}" for k,v in new_values.items())
        else:
            temp_quota = safe_int(request.form.get('temp_quota'), 'Biên chế tạm giao')
            present_count = safe_int(request.form.get('present_count'), 'Biên chế có mặt')
            assigned_quota = safe_int(request.form.get('assigned_quota'), 'Biên chế giao')
            note = (request.form.get('note') or '').strip()[:1000]
            conn.execute('''INSERT INTO submissions(department_id,temp_quota,present_count,assigned_quota,note,updated_at,updated_by)
                            VALUES (?,?,?,?,?,?,?)
                            ON CONFLICT(department_id) DO UPDATE SET temp_quota=excluded.temp_quota,
                            present_count=excluded.present_count, assigned_quota=excluded.assigned_quota,
                            note=excluded.note, updated_at=excluded.updated_at, updated_by=excluded.updated_by''',
                         (u['department_id'], temp_quota, present_count, assigned_quota, note, now, u['id']))
            new_data = f'C={temp_quota}; D={present_count}; E={assigned_quota}; H={note}'
        conn.execute('''INSERT INTO audit_logs(user_id,department_id,action,old_data,new_data,ip_address,created_at)
                        VALUES (?,?,?,?,?,?,?)''',
                     (u['id'], u['department_id'], 'Cập nhật số liệu', str(old_data), new_data, request.remote_addr, now))
        conn.commit(); conn.close()
        rebuild_current_excel()
        flash('Đã lưu dữ liệu thành công. Dữ liệu đã được cập nhật vào file Excel tổng.', 'success')
        return redirect(url_for('department_form'))
    conn.close()
    return render_template('department_form.html', sub=sub, fields=fields)


@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = db()
    units = conn.execute('''SELECT d.*, u.username, u.active AS user_active, u.id AS user_id,
                           s.updated_at, s.temp_quota, s.present_count, s.assigned_quota
                           FROM departments d
                           LEFT JOIN users u ON u.department_id=d.id
                           LEFT JOIN submissions s ON s.department_id=d.id
                           WHERE d.active=1
                           ORDER BY d.excel_row''').fetchall()
    stats = conn.execute('''SELECT (SELECT COUNT(*) FROM departments WHERE active=1) total,
                           (SELECT COUNT(*) FROM submissions s JOIN departments d ON d.id=s.department_id WHERE d.active=1) submitted''').fetchone()
    fields = get_field_settings(conn)
    conn.close()
    return render_template('admin.html', units=units, stats=stats, fields=fields)


@app.route('/admin/field-settings', methods=['POST'])
@admin_required
def update_field_settings():
    conn = db()
    fields = get_field_settings(conn)
    enabled_count = 0
    dyn_mode = bool(conn.execute('SELECT 1 FROM dynamic_fields LIMIT 1').fetchone())
    for f in fields:
        key = f['key']
        enabled = 1 if request.form.get(f'enabled_{key}') == '1' else 0
        required = 1 if enabled and request.form.get(f'required_{key}') == '1' else 0
        enabled_count += enabled
        if dyn_mode:
            conn.execute('UPDATE dynamic_fields SET enabled=?, required=? WHERE field_key=?', (enabled, required, key))
        else:
            conn.execute('''INSERT INTO field_settings(field_key,enabled,required) VALUES (?,?,?)
                            ON CONFLICT(field_key) DO UPDATE SET enabled=excluded.enabled, required=excluded.required''',
                         (key, enabled, required))
    if enabled_count == 0:
        conn.rollback(); conn.close()
        flash('Phải chọn ít nhất 1 trường cho đơn vị nhập.', 'danger')
        return redirect(url_for('admin_dashboard'))
    now = datetime.now().isoformat(timespec='seconds')
    u = current_user()
    config_text = '; '.join(
        f"{f['label']}: {'Bật' if request.form.get('enabled_'+f['key']) == '1' else 'Tắt'}"
        for f in fields
    )
    conn.execute('''INSERT INTO audit_logs(user_id,department_id,action,old_data,new_data,ip_address,created_at)
                    VALUES (?,?,?,?,?,?,?)''',
                 (u['id'], None, 'Cấu hình trường nhập liệu', None, config_text, request.remote_addr, now))
    conn.commit(); conn.close()
    flash('Đã cập nhật các trường dữ liệu cần nhập.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/excel-template', methods=['GET', 'POST'])
@admin_required
def excel_template_manager():
    conn = db()
    meta = get_template_meta(conn)
    conn.close()
    inspection = None
    pending_file = None
    selected_sheet = None
    warning = None
    if request.method == 'POST' and request.form.get('action') == 'upload':
        f = request.files.get('excel_file')
        if not f or not f.filename:
            flash('Vui lòng chọn file Excel.', 'danger')
            return redirect(url_for('excel_template_manager'))
        ext = Path(f.filename).suffix.lower()
        if ext not in ('.xlsx', '.xlsm', '.xls'):
            flash('Hỗ trợ .xlsx, .xlsm và .xls. File khác vui lòng lưu lại dưới dạng Excel trước khi tải lên.', 'danger')
            return redirect(url_for('excel_template_manager'))
        safe = secure_filename(Path(f.filename).stem) or 'excel_mau'
        temp_name = f"pending_{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe}{ext}"
        raw_path = UPLOAD_DIR / temp_name
        f.save(raw_path)
        if ext == '.xls':
            converted = raw_path.with_suffix('.xlsx')
            try:
                convert_xls_to_xlsx(raw_path, converted)
            except Exception as e:
                raw_path.unlink(missing_ok=True)
                flash(f'Không đọc được file .xls: {e}', 'danger')
                return redirect(url_for('excel_template_manager'))
            raw_path.unlink(missing_ok=True)
            raw_path = converted
            warning = 'File .xls cũ đã được chuyển sang .xlsx. Dữ liệu được giữ, nhưng định dạng hoặc công thức phức tạp có thể không giữ nguyên 100%.'
        try:
            inspection = inspect_excel(raw_path)
        except Exception as e:
            raw_path.unlink(missing_ok=True)
            flash(f'Không thể đọc file Excel: {e}', 'danger')
            return redirect(url_for('excel_template_manager'))
        pending_file = raw_path.name
        selected_sheet = inspection[0] if inspection else None
        return render_template('excel_template.html', meta=meta, inspection=inspection,
                               pending_file=pending_file, selected_sheet=selected_sheet, warning=warning)

    if request.method == 'POST' and request.form.get('action') == 'apply':
        pending = secure_filename(request.form.get('pending_file') or '')
        source = UPLOAD_DIR / pending
        if not pending or not source.exists() or source.parent != UPLOAD_DIR:
            flash('File tạm không còn tồn tại. Vui lòng tải lại file Excel.', 'danger')
            return redirect(url_for('excel_template_manager'))
        sheet_name = request.form.get('sheet_name') or ''
        try:
            header_row = int(request.form.get('header_row') or '1')
            unit_col = int(request.form.get('unit_col') or '1')
        except ValueError:
            flash('Dòng tiêu đề hoặc cột đơn vị không hợp lệ.', 'danger')
            return redirect(url_for('excel_template_manager'))
        wb = load_workbook(source, data_only=False, keep_vba=source.suffix.lower()=='.xlsm')
        if sheet_name not in wb.sheetnames:
            flash('Sheet được chọn không tồn tại.', 'danger')
            return redirect(url_for('excel_template_manager'))
        ws = wb[sheet_name]
        if header_row < 1 or header_row > ws.max_row or unit_col < 1 or unit_col > ws.max_column:
            flash('Vị trí tiêu đề/cột đơn vị nằm ngoài phạm vi file.', 'danger')
            return redirect(url_for('excel_template_manager'))

        selected_cols = []
        for c in range(1, ws.max_column + 1):
            if c == unit_col or request.form.get(f'use_col_{c}') != '1':
                continue
            label = (request.form.get(f'label_col_{c}') or '').strip()
            if not label:
                hv = ws.cell(header_row, c).value
                label = str(hv).strip() if hv not in (None, '') else f'Cột {get_column_letter(c)}'
            kind = request.form.get(f'kind_col_{c}') or 'text'
            if kind not in ('number', 'text'):
                kind = 'text'
            required = 1 if request.form.get(f'required_col_{c}') == '1' else 0
            selected_cols.append((c, label[:250], kind, required))
        if not selected_cols:
            flash('Phải chọn ít nhất 1 cột dữ liệu cho đơn vị nhập.', 'danger')
            return redirect(url_for('excel_template_manager'))

        units = []
        seen = set()
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(r, unit_col).value
            if v in (None, ''):
                continue
            name = str(v).strip()
            n = norm_text(name)
            if not name or n.startswith('tong') or n in ('cap tinh','cap xa','du phong','stt','so tt'):
                continue
            if name in seen:
                continue
            seen.add(name)
            units.append((r, name))
        if not units:
            flash('Không nhận dạng được đơn vị nào. Hãy kiểm tra cột tên đơn vị.', 'danger')
            return redirect(url_for('excel_template_manager'))

        active_ext = '.xlsm' if source.suffix.lower() == '.xlsm' else '.xlsx'
        active_path = DATA_DIR / f'active_template{active_ext}'
        if active_path.exists():
            active_path.unlink()
        source.replace(active_path)

        conn = db()
        conn.execute('UPDATE departments SET excel_row=-id, active=0')
        created_credentials = []
        for rownum, name in units:
            d = conn.execute('SELECT id FROM departments WHERE name=?', (name,)).fetchone()
            if d:
                did = d['id']
                conn.execute('UPDATE departments SET excel_row=?, active=1 WHERE id=?', (rownum, did))
            else:
                cur = conn.execute('INSERT INTO departments(name,excel_row,active) VALUES (?,?,1)', (name, rownum))
                did = cur.lastrowid
            cred = ensure_department_user(conn, did, name)
            if cred:
                created_credentials.append((name, cred[0], cred[1]))

        conn.execute('DELETE FROM dynamic_fields')
        conn.execute('DELETE FROM dynamic_values')
        conn.execute('DELETE FROM submissions')
        for order, (col, label, kind, required) in enumerate(selected_cols, start=1):
            key = f'col_{col}'
            conn.execute("""INSERT INTO dynamic_fields(field_key,label,excel_column,kind,enabled,required,sort_order)
                            VALUES (?,?,?,?,1,?,?)""", (key, label, col, kind, required, order))
        set_template_meta(conn, template_path=active_path, original_filename=request.form.get('original_filename') or pending,
                          sheet_name=sheet_name, header_row=header_row, unit_col=unit_col,
                          updated_at=datetime.now().isoformat(timespec='seconds'))

        dept_by_row = {r['excel_row']: r['id'] for r in conn.execute('SELECT id,excel_row FROM departments WHERE active=1').fetchall()}
        now = datetime.now().isoformat(timespec='seconds')
        admin = current_user()
        for rownum, _name in units:
            did = dept_by_row.get(rownum)
            for col, _label, _kind, _required in selected_cols:
                val = ws.cell(rownum, col).value
                if val is None or (isinstance(val, str) and val.startswith('=')):
                    continue
                conn.execute("""INSERT OR REPLACE INTO dynamic_values(department_id,field_key,value_text,updated_at,updated_by)
                                VALUES (?,?,?,?,?)""", (did, f'col_{col}', str(val), now, admin['id']))
        conn.execute("""INSERT INTO audit_logs(user_id,department_id,action,old_data,new_data,ip_address,created_at)
                        VALUES (?,?,?,?,?,?,?)""",
                     (admin['id'], None, 'Áp dụng file Excel mẫu mới', None,
                      f'{sheet_name}; header={header_row}; unit_col={unit_col}; units={len(units)}; fields={len(selected_cols)}',
                      request.remote_addr, now))
        conn.commit(); conn.close()
        rebuild_current_excel()

        if created_credentials:
            cred_path = DATA_DIR / 'tai_khoan_moi_sau_upload.csv'
            with open(cred_path, 'w', encoding='utf-8-sig', newline='') as fh:
                w = csv.writer(fh); w.writerow(['Đơn vị','Tên đăng nhập','Mật khẩu tạm'])
                w.writerows(created_credentials)
            flash(f'Đã áp dụng file mới: {len(units)} đơn vị, {len(selected_cols)} trường nhập. Có {len(created_credentials)} tài khoản mới.', 'success')
        else:
            flash(f'Đã áp dụng file mới: {len(units)} đơn vị và {len(selected_cols)} trường nhập. Tài khoản cũ của các đơn vị trùng tên được giữ nguyên.', 'success')
        return redirect(url_for('admin_dashboard'))

    return render_template('excel_template.html', meta=meta, inspection=None, pending_file=None,
                           selected_sheet=None, warning=None)


@app.route('/admin/unit/<int:department_id>')
@admin_required
def admin_unit(department_id):
    conn = db()
    unit = conn.execute('''SELECT d.*, u.username, u.id AS user_id, u.active AS user_active,
                          s.* FROM departments d LEFT JOIN users u ON u.department_id=d.id
                          LEFT JOIN submissions s ON s.department_id=d.id WHERE d.id=?''', (department_id,)).fetchone()
    logs = conn.execute('''SELECT a.*, u.username FROM audit_logs a LEFT JOIN users u ON u.id=a.user_id
                           WHERE a.department_id=? ORDER BY a.id DESC LIMIT 50''', (department_id,)).fetchall()
    conn.close()
    if not unit: abort(404)
    return render_template('admin_unit.html', unit=unit, logs=logs)


@app.route('/admin/reset-password/<int:user_id>', methods=['POST'])
@admin_required
def reset_password(user_id):
    new_password = secrets.token_urlsafe(7) + '!6A'
    conn = db()
    u = conn.execute("SELECT u.*, d.name AS department_name FROM users u LEFT JOIN departments d ON d.id=u.department_id WHERE u.id=? AND u.role='department'", (user_id,)).fetchone()
    if not u:
        conn.close(); abort(404)
    conn.execute('UPDATE users SET password_hash=?, must_change_password=1, failed_attempts=0, locked_until=NULL WHERE id=?', (generate_password_hash(new_password), user_id))
    conn.commit(); conn.close()
    flash(f'Mật khẩu tạm mới của {u["department_name"]}: {new_password}', 'warning')
    return redirect(url_for('admin_unit', department_id=u['department_id']))


@app.route('/admin/toggle-user/<int:user_id>', methods=['POST'])
@admin_required
def toggle_user(user_id):
    conn = db()
    u = conn.execute("SELECT * FROM users WHERE id=? AND role='department'", (user_id,)).fetchone()
    if not u: conn.close(); abort(404)
    conn.execute('UPDATE users SET active=? WHERE id=?', (0 if u['active'] else 1, user_id))
    conn.commit(); conn.close(); flash('Đã cập nhật trạng thái tài khoản.', 'success')
    return redirect(url_for('admin_unit', department_id=u['department_id']))


@app.route('/admin/credentials.csv')
@admin_required
def download_credentials():
    conn = db()
    rows = conn.execute('''SELECT d.name, u.username, u.must_change_password, u.active
                           FROM departments d JOIN users u ON u.department_id=d.id ORDER BY d.excel_row''').fetchall()
    conn.close()
    sio = StringIO(); w = csv.writer(sio)
    w.writerow(['Đơn vị','Tên đăng nhập','Mật khẩu','Ghi chú'])
    for r in rows:
        w.writerow([r['name'], r['username'], '(Không thể xem lại - hãy Đặt lại mật khẩu nếu quên)', 'Hoạt động' if r['active'] else 'Đã khóa'])
    data = BytesIO(sio.getvalue().encode('utf-8-sig'))
    return send_file(data, mimetype='text/csv', as_attachment=True, download_name='danh_sach_tai_khoan.csv')


@app.route('/admin/export.xlsx')
@admin_required
def export_excel():
    rebuild_current_excel()
    return send_file(CURRENT_XLSX, as_attachment=True, download_name='DU_LIEU_TONG_HOP.xlsx')


@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', title='Không có quyền truy cập', message='Bạn không có quyền truy cập chức năng này.'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', title='Không tìm thấy', message='Trang bạn yêu cầu không tồn tại.'), 404


init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=os.environ.get('FLASK_DEBUG') == '1')
